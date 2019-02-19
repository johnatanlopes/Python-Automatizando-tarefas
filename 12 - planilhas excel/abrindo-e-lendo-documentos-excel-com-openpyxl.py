# pip3 install openpyxl
import openpyxl, os

os.chdir('12 - planilhas excel')
wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))

# Visualizando as planihas do arquivo excel
print(wb.sheetnames)

# Verificando a planilha ativa
print(wb.active)

# Obtendo células das planilhas
sheet = wb['Sheet1']
print(sheet['A1'].value)

c = sheet['B1']
print(c.value)

print('Row ' + str(c.row) + ', Column ' + str(c.column)  + ' is ' + c.value)
print('Cell ' + str(c.coordinate) + ' is ' + c.value)

print(sheet['C1'].value)

# As linhas e colunas podem ser representadas por inteiro utilizando a função cell()
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

# Podemos determinar o tamano da planilha usando os métodos abaixo:
print(sheet.max_row) # Total de linhas
print(sheet.max_column) # Total de colunas


