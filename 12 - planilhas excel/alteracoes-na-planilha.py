'''
Definindo a altura da linha e a largura da coluna
'''

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 50
wb.save('dimensions.xlsx')


'''
Mesclar e separar celulas
'''

wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3')
sheet['A1'] = 'Twelve cells merged together'
sheet.merge_cells('C5:D5')
sheet['C5'] = 'Two merged cells'
wb.save('merged.xlsx')

'''
Removendo a mesclagem das linhas
'''

wb = openpyxl.load_workbook('merged.xlsx')
sheet = wb.active
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')
wb.save('unmerged.xlsx')

'''
Paineis congelados
'''
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2'
wb.save('freezeExample.xlsx')