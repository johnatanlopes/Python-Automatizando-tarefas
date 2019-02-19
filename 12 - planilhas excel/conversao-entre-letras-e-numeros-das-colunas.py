import openpyxl, os
from openpyxl.utils import get_column_letter, column_index_from_string

# Exibindo a coluna com numero 1
print(get_column_letter(1))

# Exibindo a coluna com numero 2
print(get_column_letter(2))

# Exibindo a coluna com numero 27
print(get_column_letter(27))

# Exibindo a coluna com numero 900
print(get_column_letter(900))

os.chdir('12 - planilhas excel')
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
print(get_column_letter(sheet.max_column))

# Exibindo o valor númerico da coluna A
print(column_index_from_string('A'))

# Exibindo o valor númerico da coluna AA
print(column_index_from_string('AA'))