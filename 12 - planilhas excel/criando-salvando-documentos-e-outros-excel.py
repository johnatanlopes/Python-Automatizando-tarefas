import openpyxl

'''
Criando workbook vazio
Exibindo o titulo
Alterando o titulo
'''
# Criando uma workbook vazio
wb = openpyxl.Workbook()
print(wb.sheetnames)

# Criando a variável sheet e passando a planilha ativa
sheet = wb.active

# Exibindo o titulo da planilha ativa
print(sheet.title)

# Alterando o titulo da planilha
sheet.title = 'Spam Bacon Eggs Sheet'
print(sheet.title)


'''
Abrindo uma planilha, alterando o titulo dela e fazendo uma copia com as alterações
'''

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'Spam Spam'
wb.save('example_copy.xlsx')


'''
Criando e removendo planilhas
'''
wb = openpyxl.Workbook()
wb.create_sheet()
print(wb.sheetnames)
wb.create_sheet(index=0, title='First Sheet')
print(wb.sheetnames)
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.sheetnames)

del wb['Sheet1']
print(wb.sheetnames)

'''
Escrevendo valores em celulas
'''
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Hello world!'
print(sheet['A1'].value)
