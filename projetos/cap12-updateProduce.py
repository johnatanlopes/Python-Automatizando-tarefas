# updateProduce.py - Corrige os preços em uma planilha de venda de produtos.

import openpyxl
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active

# Os tipos de produto e seus preços atualizados
PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

# Percorre as linhas em um loop e atualiza os preços
for rowNum in range(2, sheet.max_row): # Pula a primeira linha
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
